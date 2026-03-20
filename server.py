"""
Lovely Oven — Finance Backend Server
Run with: python server.py
Then open: http://localhost:5050/finance.html
"""

from flask import Flask, jsonify, request, send_from_directory
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
import os, datetime

app = Flask(__name__, static_folder=".")
XLSX_PATH = os.path.join(os.path.dirname(__file__), "ATOH_Finance_Database.xlsx")

# ── Helpers ───────────────────────────────────────────────────────────────────

def safe_float(val):
    if val is None:
        return 0.0
    if isinstance(val, str) and val.startswith('='):
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0

def vnd_fmt():
    return '#,##0\\ [$\u20ab-42A]'

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def load_wb():
    return load_workbook(XLSX_PATH)

def save_wb(wb):
    wb.save(XLSX_PATH)

def next_id(ws, prefix):
    existing = []
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        val = row[0]
        if val and isinstance(val, str) and val.startswith(prefix):
            try:
                existing.append(int(val[len(prefix):]))
            except ValueError:
                pass
    return f"{prefix}{(max(existing) + 1 if existing else 1):03d}"

def fmt_date(val):
    if isinstance(val, (datetime.date, datetime.datetime)):
        return val.strftime("%Y-%m-%d")
    return str(val) if val else ""

def parse_date(s):
    if not s:
        return datetime.date.today()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return datetime.date.today()

def style_row(ws, row_num, num_cols):
    for c in range(1, num_cols + 1):
        ws.cell(row_num, c).border = thin_border()
        ws.cell(row_num, c).alignment = Alignment(horizontal="left", vertical="center")

# ── CORS ──────────────────────────────────────────────────────────────────────

@app.after_request
def add_cors(response):
    response.headers["Access-Control-Allow-Origin"] = "*"
    response.headers["Access-Control-Allow-Headers"] = "Content-Type"
    response.headers["Access-Control-Allow-Methods"] = "GET,POST,PUT,DELETE,OPTIONS"
    return response

@app.route("/api/<path:path>", methods=["OPTIONS"])
def options_handler(path):
    return jsonify({}), 200

@app.route("/", defaults={"path": "finance.html"})
@app.route("/<path:path>")
def static_files(path):
    return send_from_directory(".", path)

# ═════════════════════════════════════════════════════════════════════════════
# DASHBOARD
# ═════════════════════════════════════════════════════════════════════════════

@app.route("/api/dashboard")
def dashboard():
    wb = load_wb()

    ws_ord = wb["💰 Orders"]
    total_rev = 0
    order_ids = set()
    for row in ws_ord.iter_rows(min_row=2, values_only=True):
        if row[0]:
            order_ids.add(row[0])
            # col L = index 11: line total (may be a formula string in new rows)
            # also compute from qty * price - discount as fallback
            lt = safe_float(row[11])
            if lt == 0:
                lt = safe_float(row[8]) * safe_float(row[9]) - safe_float(row[10])
            total_rev += lt

    ws_cost = wb["🧾 Costs"]
    total_cost = 0
    for row in ws_cost.iter_rows(min_row=2, values_only=True):
        if row[0]:
            total_cost += safe_float(row[7])

    return jsonify({
        "revenue": total_rev,
        "costs":   total_cost,
        "profit":  total_rev - total_cost,
        "orders":  len(order_ids),
        "margin":  (total_rev - total_cost) / total_rev if total_rev else 0,
    })

# ═════════════════════════════════════════════════════════════════════════════
# PRODUCTS
# ═════════════════════════════════════════════════════════════════════════════

@app.route("/api/products")
def get_products():
    wb = load_wb()
    ws = wb["🍜 Products"]
    products = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        cost = safe_float(row[8])
        sell = safe_float(row[9])
        products.append({
            "id":          row[0],
            "name_en":     row[1] or "",
            "name_vi":     row[2] or "",
            "category":    row[3] or "",
            "description": row[4] or "",
            "unit":        row[5] or "",
            "weight_g":    row[6] or "",
            "portion":     row[7] or "",
            "cost_price":  cost,
            "sell_price":  sell,
            "margin":      (sell - cost) / sell if sell else 0,
            "active":      row[11] or "Y",
            "notes":       row[12] or "",
        })
    return jsonify(products)

@app.route("/api/products", methods=["POST"])
def add_product():
    wb = load_wb()
    ws = wb["🍜 Products"]
    d  = request.json
    r  = ws.max_row + 1
    pid = next_id(ws, "PRD")

    ws.append([
        pid,
        d.get("name_en", ""), d.get("name_vi", ""),
        d.get("category", ""), d.get("description", ""),
        d.get("unit", ""), d.get("weight_g", ""), d.get("portion", ""),
        float(d.get("cost_price", 0)), float(d.get("sell_price", 0)),
        f"=IFERROR((J{r}-I{r})/J{r},0)",
        d.get("active", "Y"), d.get("notes", "")
    ])
    ws.cell(r, 9).number_format  = vnd_fmt()
    ws.cell(r, 9).font  = Font(name="Arial", color="0000FF", size=10)
    ws.cell(r, 10).number_format = vnd_fmt()
    ws.cell(r, 10).font = Font(name="Arial", color="0000FF", size=10)
    ws.cell(r, 11).number_format = "0.0%"
    style_row(ws, r, 13)
    save_wb(wb)
    return jsonify({"id": pid, "status": "ok"})

# ═════════════════════════════════════════════════════════════════════════════
# ORDERS
# ═════════════════════════════════════════════════════════════════════════════

@app.route("/api/orders")
def get_orders():
    wb = load_wb()
    ws = wb["💰 Orders"]
    orders = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        oid = row[0]
        qty   = safe_float(row[8])
        price = safe_float(row[9])
        disc  = safe_float(row[10])
        lt    = safe_float(row[11])
        if lt == 0:
            lt = qty * price - disc
        line = {
            "product_id":   row[4] or "",
            "product_name": row[5] or "",
            "unit":         row[6] or "",
            "weight_g":     row[7] or "",
            "qty":          qty,
            "unit_price":   price,
            "discount":     disc,
            "line_total":   lt,
        }
        if oid not in orders:
            orders[oid] = {
                "id":         oid,
                "date":       fmt_date(row[1]),
                "customer":   row[2] or "",
                "order_type": row[3] or "",
                "payment":    row[12] or "",
                "staff":      row[13] or "",
                "notes":      row[14] or "",
                "items":      [],
                "total":      0,
            }
        orders[oid]["items"].append(line)
        orders[oid]["total"] += lt

    result = sorted(orders.values(), key=lambda x: x["date"], reverse=True)
    return jsonify(result)

@app.route("/api/orders", methods=["POST"])
def add_order():
    wb = load_wb()
    ws = wb["💰 Orders"]
    d    = request.json
    oid  = next_id(ws, "ORD")
    date = parse_date(d.get("date"))

    for item in d.get("items", []):
        r    = ws.max_row + 1
        qty  = float(item.get("qty", 1))
        price= float(item.get("unit_price", 0))
        disc = float(item.get("discount", 0))
        ws.append([
            oid, date,
            d.get("customer", ""), d.get("order_type", "Dine-in"),
            item.get("product_id", ""), item.get("product_name", ""),
            item.get("unit", ""), item.get("weight_g", ""),
            qty, price, disc,
            f"=(I{r}*J{r})-K{r}",
            d.get("payment", "Cash"), d.get("staff", ""), d.get("notes", "")
        ])
        ws.cell(r, 2).number_format  = "DD/MM/YYYY"
        ws.cell(r, 10).number_format = vnd_fmt()
        ws.cell(r, 10).font = Font(name="Arial", color="0000FF", size=10)
        ws.cell(r, 11).number_format = vnd_fmt()
        ws.cell(r, 12).number_format = vnd_fmt()
        style_row(ws, r, 15)

    save_wb(wb)
    return jsonify({"id": oid, "status": "ok"})

@app.route("/api/orders/<order_id>", methods=["DELETE"])
def delete_order(order_id):
    wb = load_wb()
    ws = wb["💰 Orders"]
    rows_to_delete = [
        row[0].row for row in ws.iter_rows(min_row=2)
        if row[0].value == order_id
    ]
    for r in reversed(rows_to_delete):
        ws.delete_rows(r)
    save_wb(wb)
    return jsonify({"status": "ok", "deleted": len(rows_to_delete)})

# ═════════════════════════════════════════════════════════════════════════════
# COSTS
# ═════════════════════════════════════════════════════════════════════════════

@app.route("/api/costs")
def get_costs():
    wb = load_wb()
    ws = wb["🧾 Costs"]
    costs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        costs.append({
            "id":          row[0],
            "date":        fmt_date(row[1]),
            "supplier":    row[2] or "",
            "category":    row[3] or "",
            "item":        row[4] or "",
            "quantity":    safe_float(row[5]),
            "unit":        row[6] or "",
            "total_cost":  safe_float(row[7]),
            "receipt_ref": row[8] or "",
            "tax":         safe_float(row[9]),
            "paid_by":     row[10] or "",
            "notes":       row[11] or "",
        })
    costs.sort(key=lambda x: x["date"], reverse=True)
    return jsonify(costs)

@app.route("/api/costs", methods=["POST"])
def add_cost():
    wb  = load_wb()
    ws  = wb["🧾 Costs"]
    data = request.json
    items = data if isinstance(data, list) else [data]
    ids = []
    for item in items:
        cid  = next_id(ws, "CST")
        date = parse_date(item.get("date"))
        r    = ws.max_row + 1
        ws.append([
            cid, date,
            item.get("supplier", ""),
            item.get("category", "Other"),
            item.get("item", item.get("name", "")),
            float(item.get("quantity", 1)),
            item.get("unit", ""),
            float(item.get("total_cost", item.get("amount", 0))),
            item.get("receipt_ref", ""),
            float(item.get("tax", 0)),
            item.get("paid_by", item.get("paid", "Cash")),
            item.get("notes", "")
        ])
        ws.cell(r, 2).number_format = "DD/MM/YYYY"
        ws.cell(r, 8).number_format = vnd_fmt()
        ws.cell(r, 8).font = Font(name="Arial", color="0000FF", size=10)
        ws.cell(r, 10).number_format = vnd_fmt()
        style_row(ws, r, 12)
        ids.append(cid)
    save_wb(wb)
    return jsonify({"ids": ids, "status": "ok"})

@app.route("/api/costs/<cost_id>", methods=["DELETE"])
def delete_cost(cost_id):
    wb = load_wb()
    ws = wb["🧾 Costs"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == cost_id:
            ws.delete_rows(row[0].row)
            save_wb(wb)
            return jsonify({"status": "ok"})
    return jsonify({"status": "not_found"}), 404

# ═════════════════════════════════════════════════════════════════════════════
# INVENTORY
# ═════════════════════════════════════════════════════════════════════════════

@app.route("/api/inventory")
def get_inventory():
    wb = load_wb()
    ws = wb["📦 Inventory"]
    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        stock   = safe_float(row[5])
        min_s   = safe_float(row[6])
        reorder = safe_float(row[7])
        cpu     = safe_float(row[8])
        if stock <= min_s:
            status = "LOW"
        elif stock <= reorder:
            status = "Reorder"
        else:
            status = "OK"
        items.append({
            "id":            row[0],
            "name_en":       row[1] or "",
            "name_vi":       row[2] or "",
            "category":      row[3] or "",
            "unit":          row[4] or "",
            "stock":         stock,
            "min_stock":     min_s,
            "reorder":       reorder,
            "cost_per_unit": cpu,
            "stock_value":   stock * cpu,
            "supplier":      row[10] or "",
            "last_updated":  fmt_date(row[11]),
            "status":        status,
        })
    return jsonify(items)

# ═════════════════════════════════════════════════════════════════════════════
# MONTHLY SUMMARY
# ═════════════════════════════════════════════════════════════════════════════

@app.route("/api/monthly")
def monthly():
    wb = load_wb()
    months = {}

    ws_ord = wb["💰 Orders"]
    for row in ws_ord.iter_rows(min_row=2, values_only=True):
        if not row[0] or not row[1]:
            continue
        d = row[1]
        if isinstance(d, (datetime.date, datetime.datetime)):
            key = d.strftime("%Y-%m")
            months.setdefault(key, {"revenue": 0, "costs": 0, "orders": set()})
            lt = safe_float(row[11])
            if lt == 0:
                lt = safe_float(row[8]) * safe_float(row[9]) - safe_float(row[10])
            months[key]["revenue"] += lt
            months[key]["orders"].add(row[0])

    ws_cost = wb["🧾 Costs"]
    for row in ws_cost.iter_rows(min_row=2, values_only=True):
        if not row[0] or not row[1]:
            continue
        d = row[1]
        if isinstance(d, (datetime.date, datetime.datetime)):
            key = d.strftime("%Y-%m")
            months.setdefault(key, {"revenue": 0, "costs": 0, "orders": set()})
            months[key]["costs"] += safe_float(row[7])

    result = []
    for key in sorted(months.keys()):
        m   = months[key]
        rev = m["revenue"]
        cost= m["costs"]
        result.append({
            "month":   key,
            "revenue": rev,
            "costs":   cost,
            "profit":  rev - cost,
            "orders":  len(m["orders"]),
        })
    return jsonify(result)

# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("\n✅  Lovely Oven — Finance Server")
    print(f"   Excel file: {XLSX_PATH}")
    print("   Open in browser: http://localhost:5050/finance.html\n")
    app.run(port=5050, debug=False)
